from __future__ import annotations

from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from ..models import Chunk


def read(path: Path) -> Iterator[Chunk]:
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers: list[str | None] = []
            for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_num == 1:
                    headers = [str(c) if c is not None else None for c in row]
                    continue
                for col_idx, cell in enumerate(row):
                    if cell is None or cell == "":
                        continue
                    header = headers[col_idx] if col_idx < len(headers) else None
                    yield Chunk(
                        file_path=str(path),
                        file_type="xlsx",
                        location=f"sheet={sheet_name} row={row_num} col={header or col_idx + 1}",
                        text=str(cell),
                        column_header=header,
                    )
    finally:
        wb.close()
